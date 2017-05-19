# coding: utf-8
from openpyxl.reader.excel import load_workbook
from fwk.other.CaseDesInfo import CaseDesInfo

class CaseExcel:
    def __init__(self, path_file_excel, name_sheet=None):
        self._path_file_excel = path_file_excel
        self._name_sheet = name_sheet
        self._all_cases_info = self.getDictAllCasesInfo()

    def hasCases(self):
        if self._all_cases_info is None:
            return False
        return True

    def getDictAllCasesInfo(self):
        try:
            _WorkBook = load_workbook(filename=self._path_file_excel)
            sheet_names = _WorkBook.sheetnames
            if self._name_sheet is None:
                self._name_sheet = sheet_names[0]
            _WorkSheet = _WorkBook.get_sheet_by_name(self._name_sheet)
            max_row = _WorkSheet.max_row
            # max_column = _WorkSheet.max_column
            id_dict = {}
            tRowNum = 0
            tRowId = None
            for i in range(2, max_row + 1):
                if _WorkSheet.cell(row=i, column=CaseDesInfo.COL_ID).value is not None:
                    tRowNum = i
                    tRowId = _WorkSheet.cell(row=i, column=CaseDesInfo.COL_ID).value
                    id_dict[tRowId] = [tRowNum, i]
                else:
                    id_dict[tRowId] = [tRowNum, i]

            case_dict = {}
            for id in id_dict:
                _CaseDesInfo = CaseDesInfo()
                case_dict[id] = _CaseDesInfo
                _CaseDesInfo.id = id
                for i in range(id_dict[id][0], id_dict[id][1] + 1):
                    if _WorkSheet.cell(row=i, column=CaseDesInfo.COL_TITLE).value is not None:
                        _CaseDesInfo.title.append(_WorkSheet.cell(row=i, column=CaseDesInfo.COL_TITLE).value)
                    if _WorkSheet.cell(row=i, column=CaseDesInfo.COL_DESCRIPTION).value is not None:
                        _CaseDesInfo.description.append(
                            _WorkSheet.cell(row=i, column=CaseDesInfo.COL_DESCRIPTION).value)
                    if _WorkSheet.cell(row=i, column=CaseDesInfo.COL_EXPECTED_RESULT).value is not None:
                        _CaseDesInfo.expectedResult.append(
                            _WorkSheet.cell(row=i, column=CaseDesInfo.COL_EXPECTED_RESULT).value)
            if case_dict == {}:
                return None
            return case_dict
        except:
            return None

    def getCaseInfo(self, id):
        try:
           return self._all_cases_info[id]
        except:
            return None
