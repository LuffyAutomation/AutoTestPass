# coding: utf-8
from openpyxl.reader.excel import load_workbook
from fwk.utils.utilString.UtilString import UtilString
from fwk.utils.exceller.CaseDesInfo import CaseDesInfo


class Exceller:
    PLACEHOLDER = "placeholder"
    def __init__(self, path_file_excel, name_sheet=None):
        self._path_file_excel = path_file_excel
        self._name_sheet = name_sheet
        self.UtilString = UtilString
        # self._all_cases_info = self.getDictAllCasesInfo()

    # def hasCases(self):
    #     if self._all_cases_info is None:
    #         return False
    #     return True
    def __getWorkSheet(self):
        self._WorkBook = load_workbook(filename=self._path_file_excel, read_only=True)
        sheet_names = self._WorkBook.sheetnames
        if self._name_sheet is None:
            self._name_sheet = sheet_names[0].strip()
        else:
            self._name_sheet = self._name_sheet.strip()
        self._WorkSheet = self._WorkBook.get_sheet_by_name(self._name_sheet)

    def getDictTestDataOfAllWorkSheets(self):
        self._WorkBook = load_workbook(filename=self._path_file_excel, read_only=True)
        sheet_names = self._WorkBook.sheetnames
        testData_dict = {}
        for sheet_name in sheet_names:
            sheet_name = UtilString.toCodeName(sheet_name)
            testData_dict[sheet_name] = self.__getDictTestData(self.PLACEHOLDER, self._WorkBook.get_sheet_by_name(sheet_name))
        if testData_dict == {}:
            return None
        return testData_dict


    def __getDictTestData(self, language=PLACEHOLDER, _WorkSheet=None):
        if _WorkSheet is None:
            _WorkSheet = self._WorkSheet
        max_row = _WorkSheet.max_row
        max_column = _WorkSheet.max_column
        data_column = 0
        global value_cell
        value_cell = None
        testData_dict = {}
        for j in range(1, max_column + 1):  # 1 > ID 2 > Language
            if _WorkSheet.cell(row=3, column=j).value is not None and _WorkSheet.cell(row=3,
                                                                                                column=j).value.lower().strip() == language.lower():
                data_column = j
                break
        if data_column == 0:
            return None
            # raise Exception("Can not find [%s] column from [%s]." % (language, self._path_file_excel))

        for i in range(3, max_row + 1):
            value_cell = _WorkSheet.cell(row=i, column=1).value
            value_cell = value_cell.strip()
            if value_cell is None:
                continue
            if value_cell.lower() in ["comment", "note", "comment:", "note:", "id", "id:"]:
                continue
            value_cell = self.UtilString.toCodeName(value_cell)
            if value_cell == "":
                value_cell = "Null" + str(i)
            testData_dict[value_cell] = _WorkSheet.cell(row=i, column=data_column).value
        return testData_dict

    def getDictTestData(self, language="placeholder", _WorkSheet=None):
        try:
            self.__getWorkSheet()
            return self.__getDictTestData(language, _WorkSheet)
        except:
            return None

    def getDictAllCasesInfo(self):
        try:
            self.__getWorkSheet()
            max_row = self._WorkSheet.max_row
            # max_column = self._WorkSheet.max_column
            id_dict = {}
            tRowNum = 0
            tRowId = None
            for i in range(2, max_row + 1):
                if self._WorkSheet.cell(row=i, column=CaseDesInfo.COL_ID).value is not None:
                    tRowNum = i
                    tRowId = self._WorkSheet.cell(row=i, column=CaseDesInfo.COL_ID).value
                    id_dict[tRowId] = [tRowNum, i]
                else:
                    id_dict[tRowId] = [tRowNum, i]

            cases_dict = {}
            for id in id_dict:
                _CaseDesInfo = CaseDesInfo()
                cases_dict[id] = _CaseDesInfo
                _CaseDesInfo.id = id
                for i in range(id_dict[id][0], id_dict[id][1] + 1):
                    if self._WorkSheet.cell(row=i, column=CaseDesInfo.COL_TITLE).value is not None:
                        _CaseDesInfo.title.append(self._WorkSheet.cell(row=i, column=CaseDesInfo.COL_TITLE).value)
                    if self._WorkSheet.cell(row=i, column=CaseDesInfo.COL_DESCRIPTION).value is not None:
                        _CaseDesInfo.description.append(
                            self._WorkSheet.cell(row=i, column=CaseDesInfo.COL_DESCRIPTION).value)
                    if self._WorkSheet.cell(row=i, column=CaseDesInfo.COL_EXPECTED_RESULT).value is not None:
                        _CaseDesInfo.expectedResult.append(
                            self._WorkSheet.cell(row=i, column=CaseDesInfo.COL_EXPECTED_RESULT).value)
            if cases_dict == {}:
                return None
            return cases_dict
        except:
            return None

    def getCaseInfo(self, id):
        try:
           return self._all_cases_info[id]
        except:
            return None
